"""
Run the Negative Binomial Regression model

The regresion model use various features to predict the crime count in each
unit of study area (i.e. tract or Community Area)

Author: Hongjian
date:8/20/2015
"""

from Crime import Tract
from sets import Set
from shapely.geometry import MultiLineString
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
from scipy.stats import nbinom
from statsmodels.base.model import GenericLikelihoodModel
from openpyxl import *
import pandas as pd
from sklearn.preprocessing import scale
import statsmodels.api as sm


"""
Part One
Generate vairous features
"""


def generate_geographical_SpatialLag(foutName):
    """
    Generate the spatial lag from the geographically adjacent CAs.
    """
    fout = open(foutName, 'w')
    cnt = 0
    ts = Tract.createAllTractObjects()
    idset = Set(ts.keys())
    for i in ts.keys():
        idset.remove(i)
        for j in idset:
            if type(ts[i].polygon.intersection(ts[j].polygon)) is MultiLineString:
                fout.write('{0},{1}\n'.format(i,j))
                cnt += 1
    fout.close()
    return cnt
        
        



def generate_transition_SocialLag(year = 2010):
    """
    Generate the spatial lag from the transition flow connected CAs.
    """
    listIdx = {}
    fin = open('../data/chicago_ca_od_{0}.csv'.format(year))
    for line in fin:
        ls = line.split(",")
        srcid = int(ls[0])
        dstid = int(ls[1])
        val = int(ls[2])
        if srcid in listIdx:
            listIdx[srcid][dstid] = val
        else:
            listIdx[srcid] = {}
            listIdx[srcid][dstid] = val                            
    fin.close()

    W = np.zeros( (77,77) )
    for srcid, sdict in listIdx.items():
        total = (float) (sum( sdict.values() ))
        for dstid, val in sdict.items():
            if srcid != dstid:
                W[srcid-1][dstid-1] = val / total

    return W




def retrieve_crime_count(year, col=-1):
    """
    Retrieve the crime count in a vector
    Input:
        year - the year to retrieve
        col  - the type of crime
    """
    Y =np.zeros( (77,1) )
    with open('../data/chicago-crime-ca-level-{0}.csv'.format(year)) as fin:
        for line in fin:
            ls = line.split(",")
            idx = int(ls[0])
            val = int(ls[col])
            Y[idx-1] = val

    return Y




def retrieve_income_features():
    """
    read the xlsx file: ../data/chicago-ca-income.xlsx
    
    Three kinds of features we can generate: 
    1. population count in each category
    2. probability distribution over all categories (normalize by population)
    3. Grouped mean, variance    
    """
    wb = load_workbook("../data/chicago-ca-income.xlsx")
    ws = wb.active
    header = ws['l3':'aa3']
    header = [c.value for c in tuple(header)[0]]
    
    bins = [5000, 12500, 17500, 22500, 27500, 32500, 37500, 42500, 47500, 55000, 67500,
            87500, 112500, 137500, 175000, 300000]
    l = len(header)
    I = np.zeros((77,l))
    stats_header = ['income mean', 'std var']
    stats = np.zeros((77,2))    # mean, variance
    for idx, row in enumerate(ws.iter_rows('k4:aa80')):
        total = 0
        bin_vals = []
        for j, c in enumerate(row):
            if j == 0:
                total = float(c.value)
            else:
                I[idx][j-1] = c.value # / total
        stats[idx][0] = np.dot(bins, I[idx][:]) / total
        stats[idx][1] = np.sqrt( np.dot(I[idx][:], (bins - stats[idx][0])**2) / total )
#    return header, I
    return stats_header, stats





def retrieve_education_features():
    """
    read the xlsx file: ../data/chicago-ca-education.xlsx
    """
    wb = load_workbook("../data/chicago-ca-education.xlsx")
    ws = wb.active
    header = ws['k3':'n3']
    header = [c.value for c in tuple(header)[0]]
    
    bins = range(1,5)
    l = len(header)
    E = np.zeros((77,l))
    stats_header = ['education level', 'std var']
    stats = np.zeros((77,2))
    for i, row in enumerate(ws.iter_rows('j4:n80')):
        total = 0
        for j, c in enumerate(row):
            if j == 0:
                total = float(c.value)
            else:
                E[i][j-1] = c.value # / total
        stats[i][0] = np.dot(E[i][:], bins) / total
        stats[i][1] = np.sqrt( np.dot(E[i][:], (bins - stats[i][0])**2) / total)
    return stats_header, stats
                    
        
    
    
    
    
def retrieve_race_features():
    """
    read the xlsx file: ../data/chicago-ca-race.xlsx
    """
    wb = load_workbook("../data/chicago-ca-race.xlsx")
    ws = wb.active
    header = ws['j2':'p2']
    header = [c.value for c in tuple(header)[0]]
    l = len(header)
    R = np.zeros((77,l))
    for i, row in enumerate(ws.iter_rows('j4:p80')):
        total = 0
        for c in row:
            total += float(c.value)
        for j, c in enumerate(row):
            R[i][j] = c.value # / total
    return header, R
    
    
    


"""
Part Two
Regression models
"""

    

def linearRegression(features, Y):
    """
    learn the linear regression model from features to Y
    output the regression analysis parameters
    plot scatter plot
    """
    mod = sm.OLS(Y, features )
    res = mod.fit()
    print res.summary()


    
    

class NegBin(GenericLikelihoodModel):
    """
    negative binomial regression
    """
    
    def __init__(self, endog, exog, **kwds):
        super(NegBin, self).__init__(endog, exog, **kwds)
        
        
    def nloglikeobs(self, params):
        alpha = params[-1]
        beta = params[:-1]
        mu = np.exp(np.dot(self.exog, beta))
        size = 1 / alpha
        prob = size / (size+mu)
        ll = nbinom.logpmf( self.endog, size, prob)
        return - ll
        
    def fit(self, start_params=None, maxiter = 10000, maxfun=10000, **kwds):
        if start_params == None:
            start_params = np.append(np.zeros(self.exog.shape[1]), .5)
            if self.exog.mean() > 1:
                start_params[:-1] = np.ones(self.exog.shape[1]) * np.log(self.endog.mean())  / self.exog.mean()
            else:
                start_params[:-1] = np.ones(self.exog.shape[1]) * np.log(self.endog.mean())
            print "endog mean:", self.endog.mean(), "log endog mean:", np.log(self.endog.mean())
            print "exog mean", self.exog.mean()
        return super(NegBin, self).fit(start_params=start_params, maxiter=maxiter,
                maxfun=maxfun, **kwds)
                
        
  

    
def negativeBinomialRegression(features, Y):
    """
    learn the NB regression
    """
    mod = NegBin(Y, features)
    res = mod.fit(disp=False)
    if res.mle_retvals['converged']:
        print res.params[0], ",", res.pvalues[0]
    return res


    
def unitTest_withOnlineSource():
    import patsy
    import pandas as pd
    url = 'http://vincentarelbundock.github.com/Rdatasets/csv/COUNT/medpar.csv'
    medpar = pd.read_csv(url)
    y, X = patsy.dmatrices('los~type2+type3+hmo+white', medpar)
    res = negativeBinomialRegression(X, y)
    return y, X, medpar
    
    
def unitTest_onChicagoCrimeData():

    W = generate_transition_SocialLag(2010)
    Y = retrieve_crime_count(2010, -1)
    i = retrieve_income_features()
    e = retrieve_education_features()
    r = retrieve_race_features()
        
    f1 = np.dot(W, Y)
    f = np.concatenate((f1, i[1], e[1], r[1], np.ones(f1.shape)), axis=1)
#    f = scale(f)
#    f = np.concatenate((f, np.ones(f1.shape)), axis=1)
#    f = np.concatenate( (i[1], np.ones(f1.shape)), axis=1 )
    f = pd.DataFrame(f, columns=['social lag'] + i[0] + e[0] + r[0] + ['intercept'])
    np.savetxt("Y.csv", Y, delimiter=",")
#    f = pd.DataFrame(f, columns=i[0] + ['intercept'])
    f.to_csv("f.csv", sep=",", )
    Y = Y.reshape((77,))
    print "Y", Y.mean()
    res = negativeBinomialRegression(f, Y)
    
    
    print "f shape", f.shape
    print "Y shape", Y.shape
    linearRegression(f, Y)
    return res
     
    


def crimeRegression_eachCategory(year=2010):
    header = ['ARSON', 'ASSAULT', 'BATTERY', 'BURGLARY', 'CRIM SEXUAL ASSAULT', 
    'CRIMINAL DAMAGE', 'CRIMINAL TRESPASS', 'DECEPTIVE PRACTICE', 
    'GAMBLING', 'HOMICIDE', 'INTERFERENCE WITH PUBLIC OFFICER', 
    'INTIMIDATION', 'KIDNAPPING', 'LIQUOR LAW VIOLATION', 'MOTOR VEHICLE THEFT', 
    'NARCOTICS', 'OBSCENITY', 'OFFENSE INVOLVING CHILDREN', 'OTHER NARCOTIC VIOLATION',
    'OTHER OFFENSE', 'PROSTITUTION', 'PUBLIC INDECENCY', 'PUBLIC PEACE VIOLATION',
    'ROBBERY', 'SEX OFFENSE', 'STALKING', 'THEFT', 'WEAPONS VIOLATION', 'total']
    W = generate_transition_SocialLag(year)
    i = retrieve_income_features()
    e = retrieve_education_features()
    r = retrieve_race_features()
    predCrimes = {}
    unpredCrimes = {}
    for idx, val in enumerate(header):
        Y = retrieve_crime_count(year, idx+1)
        
        f1 = np.dot(W, Y)
        f = np.concatenate( (f1, np.ones(f1.shape)), axis=1 )
        Y = Y.reshape((77,))
        
        # linearRegression(f1, Y)
        cnt = 0
        for z in Y:
            if z == 0:
                cnt += 1
        print ",".join( [val, str(cnt), ""] ),  # sparseness
        res = negativeBinomialRegression(f, Y)
        if res.mle_retvals['converged']:
            predCrimes[val] = [cnt, len(Y)]
        else:
            unpredCrimes[val] = [cnt, len(Y)]
        
    return predCrimes, unpredCrimes
    
    
    
    
    
if __name__ == '__main__':
    # generate_geographical_SocialLag('../data/chicago-CA-geo-neighbor')
   
#   crimeRegression_eachCategory()
   f = unitTest_onChicagoCrimeData()
   print f.summary()