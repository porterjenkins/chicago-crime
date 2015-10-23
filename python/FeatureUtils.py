# -*- coding: utf-8 -*-
"""
Created on Sun Oct 04 21:37:44 2015

@author: Hongjian

Package name: featureUtil

various functions for generating feature matrices/vectors


"""

from Crime import Tract
import numpy as np
import csv
from openpyxl import *
import heapq



def generate_corina_features():
    """
    Generate the features recommended by Corina
    """
    f = open('../data/Chicago_demographics.csv', 'r')
    c = csv.reader(f)
    header = c.next()
    fields = ['totpop00_sum', 'popden00', 'pprpovW00', 'Dis46pf0', 'Stb26pf0', 'Divers5f00', 
            'pnhblWc0', 'phispWc0']
    fields_dsp = ['total population', 'population density', 'poverty index', 'disadvantage index', 'residential stability',
            'ethnic diversity', 'pct black', 'hispanic']
    hidx = []
    for fd in fields:
        hidx.append( header.index(fd) )
    
    C = np.zeros( (77,len(hidx)) )
    for i, row in enumerate(c):
        for j, k in enumerate( hidx ):
            C[i][j] = float(row[k])

    SELECTOR = [0, 2, 4, 5]
    fields_descp = [fields_dsp[i] for i in SELECTOR]
    C = C[:, SELECTOR]
    return  fields_descp, C



def generate_geographical_SpatialLag():
    """
    Generate the spatial lag from the geographically adjacent CAs.
    """
    ts = Tract.createAllTractObjects()
    ordkey = sorted(ts, key=lambda x: int(x))
    centers = [ ts[k].polygon.centroid for k in ordkey ]
    
    W = np.zeros( (len(centers), len(centers)) )
    for i, src in enumerate(centers):
        for j, dst in enumerate(centers):
            if src != dst:
                W[i][j] = src.distance(dst)
    return W
        
        

def generate_geographical_SpatialLag_ca():
    
    
    cas = Tract.createAllCAObjects()
    centers = []
    for i in range(1,78):
        centers.append(cas[i].polygon.centroid)
    
    W = np.zeros( (77,77) )
    for i, src in enumerate(centers):
        for j, dst in enumerate(centers):
            if src != dst:
                W[i][j] = src.distance(dst)
        # find n-largest (n=6)
        threshold = heapq.nlargest(6, W[i,:])[-1]
        for j in range(len(W[i,:])):
            W[i][j] = 0 if W[i][j] < threshold else W[i][j]
    return W
    
        
        


def generate_transition_SocialLag(year = 2010, lehd_type=0, region='ca'):
    """
    Generate the spatial lag from the transition flow connected CAs.
    
    0 - #jobs age under 29, 
    1 - #jobs age from 30 to 54, 
    2 - #jobs above 55, 
    3 - #jobs earning under $1250/month, 
    4 - #jobs earnings from $1251 to $3333/month, 
    5 - #jobs above $3333/month,
    6 - #jobs in goods producing, 
    7 - #jobs in trade transportation, 
    8 - #jobs in other services
    """
    
    if region == 'ca':
        ts = Tract.createAllCAObjects()
        fn = '../data/chicago_ca_od_{0}.csv'.format(year)
    elif region == 'tract':
        ts = Tract.createAllTractObjects()
        fn = '../data/chicago_od_tract_{0}.csv'.format(year)
    tsk = [int(e) for e in ts.keys()]
    ordkey = sorted(tsk)
    
    
    listIdx = {}
    fin = open(fn)
    for line in fin:
        ls = line.split(",")
        srcid = int(ls[0])
        dstid = int(ls[1])
        val = int(ls[2 + lehd_type])
        if srcid in listIdx:
            listIdx[srcid][dstid] = val
        else:
            listIdx[srcid] = {}
            listIdx[srcid][dstid] = val                            
    fin.close()
    
    W = np.zeros( (len(ts),len(ts)) )
    for srcid in ordkey:
        if srcid in listIdx:
            sdict = listIdx[srcid]
            total = (float) (sum( sdict.values() ))
            for dstid, val in sdict.items():
                if srcid != dstid:
                    if total == 0:
                        W[ordkey.index(srcid)][ordkey.index(dstid)] = 0
                    else:
                        W[ordkey.index(srcid)][ordkey.index(dstid)] = val / total
        else:
            W[ordkey.index(srcid)] = np.zeros( (1,len(ts)) )
    return W




def retrieve_crime_count(year, col=-1):
    """
    Retrieve the crime count in a vector
    Input:
        year - the year to retrieve
        col  - the type of crime
    """
    Y =np.zeros( (77,1) )
    with open('../data/chicago-crime-ca-level-{0}.csv'.format(year)) as fin:
        for line in fin:
            ls = line.split(",")
            idx = int(ls[0])
            val = int(ls[col])
            Y[idx-1] = val

    return Y




def retrieve_income_features():
    """
    read the xlsx file: ../data/chicago-ca-income.xlsx
    
    Three kinds of features we can generate: 
    1. population count in each category
    2. probability distribution over all categories (normalize by population)
    3. Grouped mean, variance    
    """
    wb = load_workbook("../data/chicago-ca-income.xlsx")
    ws = wb.active
    header = ws['l3':'aa3']
    header = [c.value for c in tuple(header)[0]]
    
#    bins = [5000, 12500, 17500, 22500, 27500, 32500, 37500, 42500, 47500, 55000, 67500,
#            87500, 112500, 137500, 175000, 300000]
    bins = range(1,17)
    l = len(header)
    I = np.zeros((77,l))
    stats_header = ['income mean', 'std var']
    stats = np.zeros((77,2))    # mean, variance
    total = np.zeros( (77,1) )
    for idx, row in enumerate(ws.iter_rows('k4:aa80')):
        bin_vals = []
        for j, c in enumerate(row):
            if j == 0:
                total[idx] =  float(c.value)
            else:
                I[idx][j-1] = c.value # / total
        stats[idx][0] = np.dot(bins, I[idx][:]) / total[idx]
        stats[idx][1] = np.sqrt( np.dot(I[idx][:], (bins - stats[idx][0])**2) / total[idx] )
#    return header, I
    return stats_header, stats, ['total'], total





def retrieve_education_features():
    """
    read the xlsx file: ../data/chicago-ca-education.xlsx
    """
    wb = load_workbook("../data/chicago-ca-education.xlsx")
    ws = wb.active
    header = ws['k3':'n3']
    header = [c.value for c in tuple(header)[0]]
    
    bins = range(1,5)
    l = len(header)
    E = np.zeros((77,l))
    stats_header = ['education level', 'std var']
    stats = np.zeros((77,2))
    for i, row in enumerate(ws.iter_rows('j4:n80')):
        total = 0
        for j, c in enumerate(row):
            if j == 0:
                total = float(c.value)
            else:
                E[i][j-1] = c.value # / total
        stats[i][0] = np.dot(E[i][:], bins) / total
        stats[i][1] = np.sqrt( np.dot(E[i][:], (bins - stats[i][0])**2) / total)
    return stats_header, stats
                    
        
    
    
    
    
def retrieve_race_features():
    """
    read the xlsx file: ../data/chicago-ca-race.xlsx
    """
    wb = load_workbook("../data/chicago-ca-race.xlsx")
    ws = wb.active
    header = ws['j2':'p2']
    header = [c.value for c in tuple(header)[0]]
    l = len(header)
    R = np.zeros((77,l))
    
    bins = range(1,8)
    
    stats_header = ['race level', 'std var']
    stats = np.zeros((77,2))
    for i, row in enumerate(ws.iter_rows('j4:p80')):
        total = 0
        for c in row:
            total += float(c.value)
        for j, c in enumerate(row):
            R[i][j] = c.value # / total
        
        stats[i][0] = np.dot(R[i][:], bins) / total
        stats[i][1] = np.sqrt( np.dot(R[i][:], (bins - stats[i][0])**2) / total)
#    return stats_header, stats
    return header, R
    
    